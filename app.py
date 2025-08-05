import os
import re
import nltk
from flask import Flask, render_template, request, flash, redirect, session, abort, jsonify
from nltk.corpus import stopwords
from nltk.stem import WordNetLemmatizer
from sklearn.feature_extraction.text import ENGLISH_STOP_WORDS
import joblib
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl

from chat import get_response
from models import Model

# Download NLTK stopwords
nltk.download('stopwords')
set(stopwords.words('english'))

# Initialize Flask app
app = Flask(__name__)
app.secret_key = os.urandom(12)

# Excel file for storing user data
EXCEL_FILE = "user_data.xlsx"
# Ensure the Excel file exists
if not os.path.exists(EXCEL_FILE):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Users"
    sheet.append(["Username", "Email", "Password"])  # Header row
    workbook.save(EXCEL_FILE)

# Preprocessing function
def preprocess_text_lemmatized(text):
    lemmatizer = WordNetLemmatizer()
    # Remove special characters and numbers
    text = re.sub(r'[^a-zA-Z\s]', '', text)
    tokens = text.lower().split()
    filtered_tokens = [lemmatizer.lemmatize(word) for word in tokens if word not in ENGLISH_STOP_WORDS]
    return ' '.join(filtered_tokens)

# Load the saved TF-IDF Vectorizer, Label Encoder, and Random Forest Model
tfidf_vectorizer_path = "saved_models/tfidf_vectorizer.joblib"
label_encoder_path = "saved_models/label_encoder.joblib"
rf_model_path = "saved_models/Random_Forest.joblib"

try:
    tfidf_vectorizer = joblib.load(tfidf_vectorizer_path)
    label_encoder = joblib.load(label_encoder_path)
    rf_model = joblib.load(rf_model_path)
    print("Models and vectorizer loaded successfully.")
except Exception as e:
    print(f"Error loading model or vectorizer: {e}")

@app.route('/')
def landing_page():
    return render_template('landing.html')

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')

        if not username or not email or not password:
            flash("All fields are required!", "danger")
            return redirect('/signup')

        # Hash the password using pbkdf2:sha256
        hashed_password = generate_password_hash(password, method='pbkdf2:sha256')

        # Save user data to Excel
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[1] == email:
                flash("Email already exists!", "danger")
                return redirect('/signup')

        sheet.append([username, email, hashed_password])
        workbook.save(EXCEL_FILE)

        flash("Signup successful! Please log in.", "success")
        return redirect('/login')

    return render_template('signup.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')

        if not email or not password:
            flash("All fields are required!", "danger")
            return redirect('/login')

        # Verify user data from Excel
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[1] == email and check_password_hash(row[2], password):
                session['user'] = row[0]
                flash("Login successful!", "success")
                return redirect('/index')

        flash("Invalid email or password!", "danger")
        return redirect('/login')

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('user', None)
    flash("You have been logged out.", "info")
    return redirect('/')

@app.route('/index')
def index_page():
    return render_template('index.html')

@app.route('/predict', methods=['POST'])
def predict():
    user_input = request.form.get('sentence')
    if not user_input:
        return render_template('sentiment.html', prediction="Please enter a sentence or use speech input.")

    try:
        # Preprocess the input and predict
        preprocessed_input = preprocess_text_lemmatized(user_input)
        X_input = tfidf_vectorizer.transform([preprocessed_input])
        model_prediction = rf_model.predict(X_input)
        decoded_prediction = label_encoder.inverse_transform(model_prediction)[0]

        # Define suggestions based on prediction
        # Make sure the string case matches what your template expects.
        if decoded_prediction == "depression":
            suggestions = [
                "Schedule a therapy session on <a href='https://www.betterhelp.com/' target='_blank'><strong>BetterHelp</strong></a>.",
                "Practice relaxation techniques using <a href='https://www.nimh.nih.gov/health/topics/relaxation-techniques' target='_blank'><strong>this guide</strong></a>.",
                "Try yoga for relaxation on <a href='https://www.youtube.com/user/yogawithadriene' target='_blank'><strong>Yoga with Adriene</strong></a>."
            ]
        elif decoded_prediction == "suicide":
            suggestions = [
                "Contact the <a href='https://988lifeline.org/' target='_blank'><strong>National Suicide Prevention Lifeline</strong></a> or call 988 immediately.",
                "Find a crisis hotline in your area via <a href='https://www.opencounseling.com/suicide-hotlines' target='_blank'><strong>Open Counseling</strong></a>.",
                "Talk to someone you trust. Learn how to open up <a href='https://www.mentalhealth.org.uk/publications/talking-mental-health' target='_blank'><strong>here</strong></a>."
            ]
        elif decoded_prediction.lower() == "non-suicide":
            suggestions = [
                "Schedule a visit with a therapist or counselor. Use BetterHelp to find professional help, Pactice deep breathing "
                "Maintain a healthy routine with tips from <a href='https://www.healthline.com/nutrition/50-super-healthy-foods' target='_blank'><strong>Healthline</strong></a>.",
                "Practice mindfulness with the <a href='https://www.headspace.com/' target='_blank'><strong>Headspace app</strong></a>.",
                "Explore hobbies for joy and relaxation on <a href='https://hobbyhelp.com/' target='_blank'><strong>Hobby Help</strong></a>."
            ]
        else:
            suggestions = []

        return render_template('sentiment.html', prediction=decoded_prediction, suggestions=suggestions)
    except Exception as e:
        flash(f"Error during prediction: {e}", "danger")
        return render_template('sentiment.html')

@app.route('/sentiment')
def sentiment():
    return render_template('sentiment.html')

@app.route('/stress_predict', methods=["POST"])
def stress_predict():
    try:
        q1 = int(request.form['a1'])
        q2 = int(request.form['a2'])
        q3 = int(request.form['a3'])
        q4 = int(request.form['a4'])
        q5 = int(request.form['a5'])
        q6 = int(request.form['a6'])
        q7 = int(request.form['a7'])
        q8 = int(request.form['a8'])
        q9 = int(request.form['a9'])
        q10 = int(request.form['a10'])
    except Exception as ex:
        flash("Invalid input for stress prediction. Please ensure all answers are numeric.", "danger")
        return redirect('/stress_predict')

    values = [q1, q2, q3, q4, q5, q6, q7, q8, q9, q10]
    model = Model()
    classifier = model.svm_classifier()
    prediction = classifier.predict([values])

    if prediction[0] == 0:
        result = 'No Depression'
        suggestions = [
            "Maintain a balanced diet. Check out <a href='https://www.healthline.com/nutrition/50-super-healthy-foods' target='_blank'><strong>healthy eating tips</strong></a>.",
            "Engage in activities you enjoy, like hobbies or sports. Learn about the benefits of hobbies <a href='https://www.psychologytoday.com/us/blog/evidence-based-living/202201/the-powerful-benefits-of-hobbies' target='_blank'><strong>here</strong></a>.",
            "Practice mindfulness with apps like <a href='https://www.headspace.com/' target='_blank'><strong>Headspace</strong></a> or <a href='https://www.calm.com/' target='_blank'><strong>Calm</strong></a>."
        ]
    elif prediction[0] == 1:
        result = 'Mild Depression'
        suggestions = [
            "Consider journaling your thoughts. Find tips <a href='https://positivepsychology.com/journaling-for-mental-health/' target='_blank'><strong>here</strong></a>.",
            "Engage in light physical activities like yoga or walking. Try beginner yoga videos on <a href='https://www.youtube.com/user/yogawithadriene' target='_blank'><strong>Yoga with Adriene</strong></a>.",
            "Talk to trusted friends or family. Learn how to open up <a href='https://www.mentalhealth.org.uk/publications/talking-mental-health' target='_blank'><strong>here</strong></a>."
        ]
    elif prediction[0] == 2:
        result = 'Moderate Depression'
        suggestions = [
            "Schedule a visit with a therapist or counselor. Use <a href='https://www.betterhelp.com/' target='_blank'><strong>BetterHelp</strong></a> to find professional help.",
            "Practice deep breathing or relaxation techniques. Check this guide on <a href='https://www.nimh.nih.gov/health/topics/relaxation-techniques' target='_blank'><strong>relaxation techniques</strong></a>.",
            "Participate in group activities or join a support group. Find one on <a href='https://mhanational.org/find-support-groups' target='_blank'><strong>Mental Health America</strong></a>."
        ]
    else:
        result = 'Severe Depression (Suicidal Risk)'
        suggestions = [
            "Seek immediate help from a mental health professional. Contact <a href='https://988lifeline.org/' target='_blank'><strong>National Suicide Prevention Lifeline</strong></a> or call <strong>988</strong>.",
            "Reach out to a crisis hotline in your area. Find international resources <a href='https://www.opencounseling.com/suicide-hotlines' target='_blank'><strong>here</strong></a>.",
            "Share your feelings with someone you trust. Learn how to talk about it <a href='https://www.samaritans.org/how-we-can-help/contact-samaritan/' target='_blank'><strong>here</strong></a>."
        ]

    return render_template("result.html", result=result, suggestions=suggestions)

@app.route('/chat', methods=['POST'])
def chat():
    user_message = request.get_json().get("message")
    if not user_message:
        return jsonify({"answer": "Please provide a message."})
    bot_response = get_response(user_message)
    return jsonify({"answer": bot_response})

if __name__ == '__main__':
    app.run(port=5987, host='0.0.0.0', debug=True)
